#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automação Sicoob v2.1 - Processamento de Extratos Bancários
Desenvolvido para processar extratos do Sicoob e gerar planilhas organizadas

Recursos:
- Template embutido protegido
- Processamento individual ou em lote
- Filtros automáticos (apenas débitos)
- Formatação preservada
- Descrições consolidadas
"""

import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re
import sys
from datetime import datetime
import shutil
from pathlib import Path

def verificar_dependencias():
    """Verifica se as bibliotecas necessárias estão instaladas"""
    bibliotecas_faltando = []
    
    try:
        import openpyxl
    except ImportError:
        bibliotecas_faltando.append("openpyxl")
    
    try:
        import chardet
    except ImportError:
        print("⚠️ Biblioteca 'chardet' não encontrada (opcional para detecção automática de encoding)")
    
    if bibliotecas_faltando:
        print("⚠️ Bibliotecas em falta:")
        for lib in bibliotecas_faltando:
            print(f"   - {lib}")
        print("\n💡 Para instalar as bibliotecas faltando, execute:")
        print(f"pip install {' '.join(bibliotecas_faltando)}")
        return False
    return True

def obter_caminho_recurso(nome_arquivo):
    """
    Obtém o caminho do arquivo de recurso (funciona tanto no desenvolvimento quanto no executável)
    """
    try:
        # Se estiver executando como executável PyInstaller
        base_path = sys._MEIPASS
    except AttributeError:
        # Se estiver executando como script Python normal
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, nome_arquivo)

def criar_planilha_usuario(nome_sugerido=None):
    """
    Cria uma nova planilha para o usuário baseada no template embutido
    """
    root = tk.Tk()
    root.withdraw()
    
    # Nome padrão baseado na sugestão ou timestamp
    if nome_sugerido:
        nome_default = nome_sugerido
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_default = f"Planilha_Sicoob_{timestamp}.xlsx"
    
    # Pergunta onde salvar a nova planilha
    caminho_destino = filedialog.asksaveasfilename(
        title="Onde salvar a planilha?",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialvalue=nome_default
    )
    
    if not caminho_destino:
        return None
    
    try:
        # Obtém o caminho do template embutido
        caminho_template = obter_caminho_recurso('Automacao_Gransoft.xlsx')
        
        if not os.path.exists(caminho_template):
            # Se não encontrar o template embutido, cria um básico
            print("⚠️ Template não encontrado. Criando planilha básica...")
            criar_planilha_basica(caminho_destino)
        else:
            # Copia o template para o local escolhido
            shutil.copy2(caminho_template, caminho_destino)
            print(f"✅ Nova planilha criada: {os.path.basename(caminho_destino)}")
        
        return caminho_destino
        
    except Exception as e:
        print(f"❌ Erro ao criar planilha: {e}")
        return None

def criar_planilha_basica(caminho_destino):
    """
    Cria uma planilha básica caso o template não esteja disponível
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Cria novo workbook
        wb = Workbook()
        
        # Aba Banco
        ws_banco = wb.active
        ws_banco.title = "Banco"
        
        # Cabeçalhos
        cabecalhos = [
            "Data Vencimento", "Descrição", "Valor", "Fornecedor", 
            "Numero Docto", "Conta Contábil", "Observação (opcional)"
        ]
        
        # Adiciona cabeçalhos com formatação
        for col, header in enumerate(cabecalhos, 1):
            cell = ws_banco.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Ajusta largura das colunas
        ws_banco.column_dimensions['A'].width = 15  # Data
        ws_banco.column_dimensions['B'].width = 50  # Descrição
        ws_banco.column_dimensions['C'].width = 15  # Valor
        ws_banco.column_dimensions['D'].width = 25  # Fornecedor
        ws_banco.column_dimensions['E'].width = 15  # Documento
        ws_banco.column_dimensions['F'].width = 20  # Conta Contábil
        ws_banco.column_dimensions['G'].width = 30  # Observação
        
        # Cria aba Base de dados
        ws_base = wb.create_sheet("Base de dados")
        ws_base.cell(row=1, column=1, value="Nome")
        ws_base.cell(row=1, column=3, value="Conta Contábil") 
        ws_base.cell(row=1, column=5, value="Colaboradores, prestadores, funcionário E FORNECEDORES")
        
        # Salva a planilha
        wb.save(caminho_destino)
        print("✅ Planilha básica criada com sucesso!")
        
    except Exception as e:
        print(f"❌ Erro ao criar planilha básica: {e}")
        raise

def processar_formato_valor_sicoob(valor_str):
    """
    Processa o formato específico do Sicoob: '- 125,69 D' ou '2.794,76 C'
    Retorna valor numérico (positivo para débitos, None para créditos)
    """
    if not valor_str or pd.isna(valor_str):
        return None
    
    valor_str = str(valor_str).strip()
    
    # Ignora valores vazios
    if not valor_str or valor_str == "nan":
        return None
    
    # Remove espaços extras
    valor_str = re.sub(r'\s+', ' ', valor_str)
    
    # Padrão para formato Sicoob: opcionalmente "- " seguido de número com vírgula e " D" ou " C"
    # Exemplos: "- 125,69 D", "2.794,76 C", "- 2.460,73 D"
    padrao = r'^-?\s*(\d{1,3}(?:\.\d{3})*,\d{2})\s*([DC])$'
    match = re.match(padrao, valor_str)
    
    if match:
        numero_str = match.group(1)  # Ex: "125,69" ou "2.460,73"
        tipo = match.group(2)        # "D" para débito ou "C" para crédito
        
        # Converte para float (troca vírgula por ponto, remove pontos dos milhares)
        numero_str = numero_str.replace('.', '').replace(',', '.')
        valor_numerico = float(numero_str)
        
        # Se for débito (D), retorna positivo (pois vamos remover o sinal de menos depois)
        # Se for crédito (C), retorna None (pois não queremos incluir)
        if tipo == 'D':
            return valor_numerico
        elif tipo == 'C':
            return None  # Créditos são explicitamente ignorados
    
    # Se não conseguiu processar no formato Sicoob, tenta formato genérico
    try:
        # Verifica se tem indicador de crédito
        if 'C' in valor_str.upper():
            return None  # Ignora créditos
        
        # Remove tudo que não é dígito, vírgula, ponto ou sinal de menos
        valor_limpo = re.sub(r'[^\d.,-]', '', valor_str)
        if valor_limpo:
            valor_limpo = valor_limpo.replace(',', '.')
            valor_numerico = float(valor_limpo)
            # Se tinha sinal de menos no original ou não tem indicador de crédito, considera como débito
            if '-' in valor_str or 'D' in valor_str.upper():
                return valor_numerico
    except:
        pass
    
    return None

def adicionar_dados_preservando_formatacao(caminho_planilha, novos_dados):
    """
    Adiciona novos dados à planilha preservando toda a formatação original
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import NamedStyle
        from openpyxl.utils import get_column_letter
        
        print("🎨 Carregando planilha preservando formatação...")
        
        # Carrega a planilha mantendo formatação
        wb = load_workbook(caminho_planilha)
        ws = wb['Banco']
        
        # Encontra a primeira linha vazia (após os dados existentes)
        linha_inicio = 1
        while linha_inicio <= ws.max_row:
            # Verifica se a linha está completamente vazia
            linha_vazia = True
            for col in range(1, 8):  # Colunas A até G
                cell_value = ws.cell(row=linha_inicio, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    linha_vazia = False
                    break
            
            if linha_vazia and linha_inicio > 1:  # Não conta a linha de cabeçalho
                break
            linha_inicio += 1
        
        print(f"📍 Iniciando inserção na linha {linha_inicio}")
        
        # Copia a formatação da linha de cabeçalho ou da última linha com dados
        linha_formato_referencia = 1 if linha_inicio <= 2 else linha_inicio - 1
        
        # Adiciona os novos dados
        for i, (index, row) in enumerate(novos_dados.iterrows()):
            linha_atual = linha_inicio + i
            
            # Data Vencimento (Coluna A)
            cell_data = ws.cell(row=linha_atual, column=1)
            try:
                # Converte string para data se necessário
                if isinstance(row['Data Vencimento'], str):
                    data_obj = datetime.strptime(row['Data Vencimento'], '%d/%m/%Y')
                    cell_data.value = data_obj
                else:
                    cell_data.value = row['Data Vencimento']
                # Aplica formatação de data
                cell_data.number_format = 'DD/MM/YYYY'
            except:
                cell_data.value = row['Data Vencimento']
            
            # Descrição (Coluna B)
            ws.cell(row=linha_atual, column=2, value=row['Descrição'])
            
            # Valor (Coluna C) - com formatação de moeda brasileira
            cell_valor = ws.cell(row=linha_atual, column=3)
            try:
                cell_valor.value = float(row['Valor'])
                cell_valor.number_format = 'R$ #,##0.00'
            except:
                cell_valor.value = row['Valor']
            
            # Fornecedor (Coluna D)
            ws.cell(row=linha_atual, column=4, value=row['Fornecedor'])
            
            # Numero Docto (Coluna E)
            ws.cell(row=linha_atual, column=5, value=row['Numero Docto'])
            
            # Conta Contábil (Coluna F)
            ws.cell(row=linha_atual, column=6, value=row['Conta Contábil'])
            
            # Observação (Coluna G)
            ws.cell(row=linha_atual, column=7, value=row['Observação (opcional)'])
            
            # Copia formatação da linha de referência (borda, alinhamento, etc.)
            if linha_formato_referencia > 0:
                for col in range(1, 8):
                    cell_origem = ws.cell(row=linha_formato_referencia, column=col)
                    cell_destino = ws.cell(row=linha_atual, column=col)
                    
                    # Copia formatação (exceto número que já definimos)
                    if cell_origem.font:
                        cell_destino.font = cell_origem.font
                    if cell_origem.border:
                        cell_destino.border = cell_origem.border
                    if cell_origem.fill:
                        cell_destino.fill = cell_origem.fill
                    if cell_origem.alignment:
                        cell_destino.alignment = cell_origem.alignment
        
        # Ajusta largura das colunas se necessário
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo de 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salva a planilha
        wb.save(caminho_planilha)
        print(f"✅ Dados adicionados preservando formatação original!")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro ao preservar formatação: {e}")
        return False

def processar_extrato_individual(caminho_extrato, caminho_planilha_usuario, mostrar_detalhes=True):
    """
    Processa um único extrato e adiciona à planilha especificada
    """
    try:
        if mostrar_detalhes:
            print(f"🔄 Processando: {os.path.basename(caminho_extrato)}")

        # --- 1. LEITURA ROBUSTA DO EXTRATO ---
        df_extrato = None
        
        # Primeiro tenta ler como Excel
        try:
            df_extrato = pd.read_excel(caminho_extrato, skiprows=1, header=0)
            if mostrar_detalhes:
                print("✅ Arquivo lido como Excel")
        except ImportError:
            if mostrar_detalhes:
                print("⚠️ Biblioteca 'openpyxl' não está instalada. Tentando como CSV...")
        except Exception as excel_error:
            if mostrar_detalhes:
                print(f"⚠️ Erro ao ler como Excel: {excel_error}")
            
        # Se não conseguiu ler como Excel, tenta CSV com diferentes encodings
        if df_extrato is None:
            csv_lido = False
            estrategias_csv = [
                {'sep': ',', 'quotechar': '"', 'encoding': 'windows-1252'},
                {'sep': ',', 'quotechar': '"', 'encoding': 'cp1252'},
                {'sep': ',', 'quotechar': '"', 'encoding': 'iso-8859-1'},
                {'sep': ';', 'quotechar': '"', 'encoding': 'windows-1252'},
                {'sep': ',', 'quotechar': '"', 'encoding': 'utf-8-sig'},
                {'sep': ',', 'quotechar': '"', 'encoding': None},
                {'sep': ',', 'quotechar': '"', 'encoding': 'utf-8'},
            ]
            
            for i, estrategia in enumerate(estrategias_csv):
                try:
                    if mostrar_detalhes:
                        print(f"🔄 Tentando estratégia CSV {i+1} (encoding: {estrategia.get('encoding', 'auto')})...")
                    
                    # Se encoding for None, tenta detectar automaticamente
                    if estrategia['encoding'] is None:
                        try:
                            import chardet
                            with open(caminho_extrato, 'rb') as f:
                                raw_data = f.read()
                                result = chardet.detect(raw_data)
                                detected_encoding = result['encoding']
                                estrategia['encoding'] = detected_encoding
                                if mostrar_detalhes:
                                    print(f"🔍 Encoding detectado: {detected_encoding}")
                        except ImportError:
                            if mostrar_detalhes:
                                print("⚠️ Biblioteca 'chardet' não disponível, usando fallback")
                            estrategia['encoding'] = 'windows-1252'
                        except Exception:
                            estrategia['encoding'] = 'windows-1252'
                    
                    df_extrato = pd.read_csv(
                        caminho_extrato, 
                        skiprows=1, 
                        header=None,
                        on_bad_lines='skip',  # Pula linhas problemáticas
                        engine='python',      # Mais flexível
                        **estrategia
                    )
                    
                    # Verifica se conseguiu ler dados válidos
                    if not df_extrato.empty and df_extrato.shape[1] >= 3:
                        if mostrar_detalhes:
                            print(f"✅ Arquivo lido como CSV com estratégia {i+1} (encoding: {estrategia['encoding']})")
                        csv_lido = True
                        break
                        
                except Exception as csv_error:
                    if mostrar_detalhes:
                        print(f"❌ Estratégia CSV {i+1} falhou: {csv_error}")
                    continue
            
            if not csv_lido:
                raise Exception(
                    "Não foi possível ler o arquivo em nenhum formato suportado.\n\n"
                    "Soluções possíveis:\n"
                    "1. Certifique-se que o arquivo é um extrato válido do Sicoob\n"
                    "2. Tente salvar o extrato em formato .xlsx (Excel)\n"
                    "3. Verifique se o arquivo não está corrompido"
                )
            
        # Lê apenas os dados existentes da planilha do usuário para comparação
        df_banco = pd.read_excel(caminho_planilha_usuario, sheet_name='Banco', engine='openpyxl')
        
        # --- 2. PREPARO DOS DADOS DO EXTRATO ---
        colunas_necessarias = ['DATA', 'DOCUMENTO', 'HISTORICO', 'VALOR']
        
        # Define as colunas baseado no que foi lido
        if df_extrato.shape[1] >= 4:
            df_extrato.columns = ['DATA', 'DOCUMENTO', 'HISTORICO', 'VALOR'] + [f'EXTRA_{i}' for i in range(df_extrato.shape[1] - 4)]
        elif df_extrato.shape[1] == 3:
            df_extrato.columns = ['DATA', 'HISTORICO', 'VALOR']
            df_extrato['DOCUMENTO'] = ''
        else:
            raise Exception(f"Arquivo tem estrutura inesperada com {df_extrato.shape[1]} colunas")
        
        for coluna in colunas_necessarias:
            if coluna not in df_extrato.columns:
                df_extrato[coluna] = ''
        
        # Seleciona apenas as colunas necessárias
        df_extrato = df_extrato[colunas_necessarias].copy()
        
        # Validação básica dos dados
        if df_extrato.empty:
            raise Exception("O arquivo está vazio ou não contém dados válidos")
        
        df_extrato = df_extrato.dropna(how='all')
        
        if len(df_extrato) == 0:
            raise Exception("Não foram encontrados dados válidos no arquivo após limpeza")
        
        # --- 3. CONSOLIDANDO DESCRIÇÕES ---
        registros_consolidados = []
        historico_atual = ""
        linha_principal = None
        transacoes_processadas = 0
        linhas_credito_ignoradas = 0
        ignorando_continuacoes_credito = False
        
        for index, row in df_extrato.iterrows():
            data_str = str(row['DATA']).strip() if pd.notna(row['DATA']) else ""
            historico_str = str(row['HISTORICO']).strip() if pd.notna(row['HISTORICO']) else ""
            valor_str = str(row['VALOR']).strip() if pd.notna(row['VALOR']) else ""
            
            # Se tem DATA válida, pode ser uma linha principal de transação
            if data_str and data_str != "" and data_str != "nan":
                ignorando_continuacoes_credito = False
                
                # Verifica se é crédito ou saldo
                eh_credito = False
                eh_saldo = False
                
                if valor_str and ("C" in valor_str or "c" in valor_str.lower()):
                    eh_credito = True
                
                frases_saldo = ['SALDO DO DIA', 'SALDO ANTERIOR', 'SALDO ATUAL', 'SALDO FINAL']
                for frase in frases_saldo:
                    if frase.lower() in historico_str.lower():
                        eh_saldo = True
                        break
                
                # Se for crédito ou saldo, ignora esta linha E suas continuações
                if eh_credito or eh_saldo:
                    linhas_credito_ignoradas += 1
                    ignorando_continuacoes_credito = True
                    continue
                
                # Se já tínhamos uma linha principal anterior, salva ela
                if linha_principal is not None:
                    linha_principal['HISTORICO'] = historico_atual.strip()
                    registros_consolidados.append(linha_principal.copy())
                    transacoes_processadas += 1
                
                # Inicia uma nova linha principal
                linha_principal = row.copy()
                historico_atual = historico_str
                
            # Se não tem DATA, é uma linha de continuação da descrição
            elif historico_str and historico_str != "":
                # Se estamos ignorando continuações de crédito, pula esta linha
                if ignorando_continuacoes_credito:
                    continue
                
                # Se temos uma linha principal válida (débito), adiciona a continuação
                if linha_principal is not None:
                    # Verifica se esta continuação não é uma linha de saldo
                    eh_continuacao_saldo = False
                    frases_saldo = ['SALDO DO DIA', 'SALDO ANTERIOR', 'SALDO ATUAL', 'SALDO FINAL']
                    for frase in frases_saldo:
                        if frase.lower() in historico_str.lower():
                            eh_continuacao_saldo = True
                            break
                    
                    if not eh_continuacao_saldo:
                        if historico_atual:
                            historico_atual += " " + historico_str
                        else:
                            historico_atual = historico_str
        
        # Adiciona a última linha
        if linha_principal is not None:
            linha_principal['HISTORICO'] = historico_atual.strip()
            registros_consolidados.append(linha_principal.copy())
            transacoes_processadas += 1
        
        # Converte a lista de volta para DataFrame
        df_extrato_consolidado = pd.DataFrame(registros_consolidados)
        
        if df_extrato_consolidado.empty:
            if mostrar_detalhes:
                print("⚠️ Nenhuma transação de débito foi encontrada após consolidação")
            return {
                'sucesso': True,
                'transacoes_processadas': 0,
                'debitos_encontrados': 0,
                'novos_lancamentos': 0,
                'duplicatas_ignoradas': 0
            }

        # --- 4. FILTROS ADICIONAIS ---
        frases_a_ignorar = [
            'SALDO DO DIA', 'SALDO ANTERIOR', 'SALDO ATUAL', 'SALDO FINAL',
            'saldo do dia', 'saldo anterior', 'saldo atual', 'saldo final',
            'Saldo bloqueado anterior', 'Saldo bloqueado', 'Saldo disponível', 'Saldo em conta'
        ]
        
        linhas_antes_filtro = len(df_extrato_consolidado)
        
        for frase in frases_a_ignorar:
            df_extrato_consolidado = df_extrato_consolidado[
                ~df_extrato_consolidado['HISTORICO'].str.contains(frase, case=False, na=False)
            ]
        
        linhas_filtradas = linhas_antes_filtro - len(df_extrato_consolidado)

        # --- 5. PROCESSAMENTO DOS VALORES ---
        df_extrato_consolidado = df_extrato_consolidado.dropna(subset=['DATA'])
        df_extrato_consolidado['HISTORICO'] = df_extrato_consolidado['HISTORICO'].str.replace(r'\s+', ' ', regex=True).str.strip()
        
        valores_processados = []
        valores_validos = 0
        valores_credito_ignorados = 0
        
        for index, row in df_extrato_consolidado.iterrows():
            valor_original = row['VALOR']
            valor_processado = processar_formato_valor_sicoob(valor_original)
            valores_processados.append(valor_processado)
            
            if valor_processado is not None:
                valores_validos += 1
            elif str(valor_original).strip() and 'C' in str(valor_original).upper():
                valores_credito_ignorados += 1
        
        df_extrato_consolidado['VALOR_PROCESSADO'] = valores_processados
        
        # --- 6. FILTRAR APENAS VALORES DE DÉBITO VÁLIDOS ---
        df_extrato_debitos = df_extrato_consolidado[df_extrato_consolidado['VALOR_PROCESSADO'].notna()].copy()
        
        if df_extrato_debitos.empty:
            if mostrar_detalhes:
                print("⚠️ Nenhuma transação de débito válida foi encontrada no extrato.")
            return {
                'sucesso': True,
                'transacoes_processadas': transacoes_processadas,
                'debitos_encontrados': 0,
                'novos_lancamentos': 0,
                'duplicatas_ignoradas': 0
            }
        
        # --- 7. MAPEAMENTO PARA A ESTRUTURA DA PLANILHA ---
        novos_lancamentos = pd.DataFrame({
            'Data Vencimento': df_extrato_debitos['DATA'],
            'Descrição': df_extrato_debitos['HISTORICO'],
            'Valor': df_extrato_debitos['VALOR_PROCESSADO'],
            'Fornecedor': '',
            'Numero Docto': df_extrato_debitos['DOCUMENTO'],
            'Conta Contábil': '',
            'Observação (opcional)': df_extrato_debitos['HISTORICO']
        })

        # --- 8. PREVENÇÃO DE DUPLICIDADE ---
        colunas_para_comparar = ['Data Vencimento', 'Descrição', 'Valor']
        df_banco_temp = df_banco.dropna(subset=colunas_para_comparar).copy()
        novos_lancamentos_temp = novos_lancamentos.dropna(subset=colunas_para_comparar).copy()
        
        # Normalizar datas para comparação
        try:
            df_banco_temp['Data Vencimento'] = pd.to_datetime(df_banco_temp['Data Vencimento'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
            novos_lancamentos_temp['Data Vencimento'] = pd.to_datetime(novos_lancamentos_temp['Data Vencimento'], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y')
        except:
            pass
        
        # Criar ID único para comparação
        df_banco_temp['ID'] = (
            df_banco_temp['Data Vencimento'].astype(str).str.strip() + '|' +
            df_banco_temp['Descrição'].astype(str).str.strip().str.lower() + '|' +
            df_banco_temp['Valor'].astype(str)
        )

        novos_lancamentos_temp['ID'] = (
            novos_lancamentos_temp['Data Vencimento'].astype(str).str.strip() + '|' +
            novos_lancamentos_temp['Descrição'].astype(str).str.strip().str.lower() + '|' +
            novos_lancamentos_temp['Valor'].astype(str)
        )

        novos_lancamentos_sem_duplicatas = novos_lancamentos[~novos_lancamentos_temp['ID'].isin(df_banco_temp['ID'])].copy()
        
        duplicatas_encontradas = len(novos_lancamentos) - len(novos_lancamentos_sem_duplicatas)
        
        if novos_lancamentos_sem_duplicatas.empty:
            if mostrar_detalhes:
                print("ℹ️ Não há novas transações para adicionar. Todas já existem na planilha.")
            return {
                'sucesso': True,
                'transacoes_processadas': transacoes_processadas,
                'debitos_encontrados': len(df_extrato_debitos),
                'novos_lancamentos': 0,
                'duplicatas_ignoradas': duplicatas_encontradas
            }
            
        # --- 9. ADICIONANDO OS DADOS PRESERVANDO FORMATAÇÃO ---
        sucesso_formatacao = adicionar_dados_preservando_formatacao(
            caminho_planilha_usuario, 
            novos_lancamentos_sem_duplicatas
        )
        
        if not sucesso_formatacao:
            # Fallback: usar pandas se falhar a preservação de formatação
            if mostrar_detalhes:
                print("⚠️ Fallback: usando método padrão sem preservar formatação completa")
            all_sheets = pd.read_excel(caminho_planilha_usuario, sheet_name=None, engine='openpyxl')
            df_banco_atualizado = pd.concat([df_banco, novos_lancamentos_sem_duplicatas], ignore_index=True)

            # Salvar na planilha preservando outras abas
            with pd.ExcelWriter(caminho_planilha_usuario, engine='openpyxl', mode='w') as writer:
                df_banco_atualizado.to_excel(writer, sheet_name='Banco', index=False)
                for sheet_name, df_sheet in all_sheets.items():
                    if sheet_name != 'Banco':
                        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            'sucesso': True,
            'transacoes_processadas': transacoes_processadas,
            'debitos_encontrados': len(df_extrato_debitos),
            'novos_lancamentos': len(novos_lancamentos_sem_duplicatas),
            'duplicatas_ignoradas': duplicatas_encontradas
        }
        
    except Exception as e:
        return {
            'sucesso': False,
            'erro': str(e),
            'transacoes_processadas': 0,
            'debitos_encontrados': 0,
            'novos_lancamentos': 0,
            'duplicatas_ignoradas': 0
        }

def processar_multiplos_extratos():
    """
    Processa múltiplos extratos, cada um em uma planilha separada
    """
    root = tk.Tk()
    root.withdraw()
    
    print("📁 Processamento em Lote - Múltiplos Extratos")
    print("=" * 60)

    # Seleciona múltiplos arquivos de extrato
    caminhos_extratos = filedialog.askopenfilenames(
        title="Selecione os extratos do Sicoob (múltiplos arquivos)",
        filetypes=[("Arquivos Excel/CSV", "*.xlsx;*.xls;*.csv")]
    )
    
    if not caminhos_extratos:
        print("❌ Operação cancelada. Nenhum arquivo de extrato foi selecionado.")
        return False
    
    # Pergunta onde salvar as planilhas
    pasta_destino = filedialog.askdirectory(
        title="Escolha a pasta onde salvar as planilhas processadas"
    )
    
    if not pasta_destino:
        print("❌ Operação cancelada. Nenhuma pasta de destino foi selecionada.")
        return False
    
    print(f"📂 Pasta de destino: {pasta_destino}")
    print(f"📄 {len(caminhos_extratos)} arquivo(s) selecionado(s) para processamento")
    print()
    
    # Cria janela de progresso
    janela_progresso = tk.Toplevel()
    janela_progresso.title("🔄 Processando Múltiplos Extratos")
    janela_progresso.geometry("500x300")
    janela_progresso.resizable(False, False)
    
    # Centralizar janela
    janela_progresso.transient()
    janela_progresso.grab_set()
    
    # Frame principal
    frame_principal = tk.Frame(janela_progresso, padx=20, pady=20)
    frame_principal.pack(fill=tk.BOTH, expand=True)
    
    # Label título
    label_titulo = tk.Label(frame_principal, text="📊 Processamento em Lote", 
                           font=('Arial', 14, 'bold'), fg='#2E7D32')
    label_titulo.pack(pady=(0, 20))
    
    # Label status
    label_status = tk.Label(frame_principal, text="Iniciando processamento...", 
                           font=('Arial', 11))
    label_status.pack(pady=5)
    
    # Barra de progresso
    progress_bar = ttk.Progressbar(frame_principal, length=400, mode='determinate')
    progress_bar.pack(pady=10)
    progress_bar['maximum'] = len(caminhos_extratos)
    
    # Label arquivo atual
    label_arquivo = tk.Label(frame_principal, text="", 
                            font=('Arial', 9), fg='gray')
    label_arquivo.pack(pady=5)
    
    # Área de texto para resultados
    texto_resultados = tk.Text(frame_principal, height=8, width=60, 
                              font=('Courier', 9))
    texto_resultados.pack(pady=10, fill=tk.BOTH, expand=True)
    
    # Scrollbar para a área de texto
    scrollbar = tk.Scrollbar(texto_resultados)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    texto_resultados.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=texto_resultados.yview)
    
    janela_progresso.update()
    
    # Processa cada extrato
    resultados_processamento = []
    sucessos = 0
    falhas = 0
    
    for i, caminho_extrato in enumerate(caminhos_extratos):
        nome_arquivo = os.path.basename(caminho_extrato)
        nome_sem_ext = os.path.splitext(nome_arquivo)[0]
        
        # Atualiza interface
        progress_bar['value'] = i
        label_status.config(text=f"Processando {i+1}/{len(caminhos_extratos)} arquivos...")
        label_arquivo.config(text=f"📄 {nome_arquivo}")
        janela_progresso.update()
        
        # Gera nome da planilha baseado no nome do extrato
        nome_planilha = f"Processado_{nome_sem_ext}.xlsx"
        caminho_planilha = os.path.join(pasta_destino, nome_planilha)
        
        # Cria planilha para este extrato
        try:
            # Cria planilha nova baseada no template
            if criar_planilha_usuario.__code__.co_argcount > 0:
                # Passa o nome sugerido para a função
                caminho_planilha_criada = criar_nova_planilha_silenciosa(caminho_planilha)
            else:
                caminho_planilha_criada = caminho_planilha
                # Copia template para o destino
                caminho_template = obter_caminho_recurso('Automacao_Gransoft.xlsx')
                if os.path.exists(caminho_template):
                    shutil.copy2(caminho_template, caminho_planilha_criada)
                else:
                    criar_planilha_basica(caminho_planilha_criada)
            
            # Processa o extrato
            resultado = processar_extrato_individual(
                caminho_extrato, 
                caminho_planilha_criada, 
                mostrar_detalhes=False
            )
            
            if resultado['sucesso']:
                sucessos += 1
                status = "✅ SUCESSO"
                msg = f"{status} - {nome_arquivo}\n"
                msg += f"   📊 {resultado['novos_lancamentos']} novos lançamentos adicionados\n"
                msg += f"   💰 {resultado['debitos_encontrados']} débitos encontrados\n"
                if resultado['duplicatas_ignoradas'] > 0:
                    msg += f"   🔍 {resultado['duplicatas_ignoradas']} duplicatas ignoradas\n"
                msg += f"   💾 Salvo em: {nome_planilha}\n\n"
            else:
                falhas += 1
                status = "❌ FALHA"
                msg = f"{status} - {nome_arquivo}\n"
                msg += f"   🚫 Erro: {resultado.get('erro', 'Erro desconhecido')}\n\n"
            
            # Adiciona resultado à interface
            texto_resultados.insert(tk.END, msg)
            texto_resultados.see(tk.END)
            janela_progresso.update()
            
            resultados_processamento.append({
                'arquivo': nome_arquivo,
                'planilha': nome_planilha,
                'resultado': resultado
            })
            
        except Exception as e:
            falhas += 1
            msg = f"❌ FALHA - {nome_arquivo}\n"
            msg += f"   🚫 Erro: {str(e)}\n\n"
            texto_resultados.insert(tk.END, msg)
            texto_resultados.see(tk.END)
            janela_progresso.update()
    
    # Finaliza progresso
    progress_bar['value'] = len(caminhos_extratos)
    label_status.config(text="Processamento concluído!")
    label_arquivo.config(text="")
    
    # Resumo final
    resumo = f"\n{'='*50}\n"
    resumo += f"📊 RESUMO FINAL:\n"
    resumo += f"✅ Sucessos: {sucessos}\n"
    resumo += f"❌ Falhas: {falhas}\n"
    resumo += f"📁 Planilhas criadas em: {pasta_destino}\n"
    resumo += f"{'='*50}\n"
    
    texto_resultados.insert(tk.END, resumo)
    texto_resultados.see(tk.END)
    
    # Botão para fechar e abrir pasta
    frame_botoes = tk.Frame(frame_principal)
    frame_botoes.pack(pady=10, fill=tk.X)
    
    def abrir_pasta_e_fechar():
        try:
            if os.name == 'nt':  # Windows
                os.startfile(pasta_destino)
            else:  # Linux/Mac
                import subprocess
                subprocess.call(['xdg-open', pasta_destino])
        except Exception as e:
            print(f"Erro ao abrir pasta: {e}")
        janela_progresso.destroy()
    
    tk.Button(frame_botoes, text="📁 Abrir Pasta", command=abrir_pasta_e_fechar,
              bg='#4CAF50', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=5)
    
    tk.Button(frame_botoes, text="✅ Fechar", command=janela_progresso.destroy,
              bg='#2196F3', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=5)
    
    print(f"\n📊 Processamento em lote concluído!")
    print(f"✅ Sucessos: {sucessos}")
    print(f"❌ Falhas: {falhas}")
    print(f"📁 Planilhas salvas em: {pasta_destino}")
    
    return True

def criar_nova_planilha_silenciosa(caminho_destino):
    """
    Cria uma nova planilha silenciosamente (sem diálogos)
    """
    try:
        # Obtém o caminho do template embutido
        caminho_template = obter_caminho_recurso('Automacao_Gransoft.xlsx')
        
        if not os.path.exists(caminho_template):
            # Se não encontrar o template embutido, cria um básico
            criar_planilha_basica(caminho_destino)
        else:
            # Copia o template para o local escolhido
            shutil.copy2(caminho_template, caminho_destino)
        
        return caminho_destino
        
    except Exception as e:
        print(f"❌ Erro ao criar planilha: {e}")
        raise

def processar_extrato_unico():
    """
    Processa um único extrato (modo tradicional)
    """
    root = tk.Tk()
    root.withdraw()

    print("📄 Processamento Individual - Extrato Único")
    print("=" * 60)

    # Passo 1: Criar nova planilha para o usuário
    print("📁 Primeiro, vamos criar sua planilha de controle...")
    caminho_planilha_usuario = criar_planilha_usuario()
    
    if not caminho_planilha_usuario:
        messagebox.showinfo("Cancelado", "Operação cancelada. Nenhuma planilha foi criada.")
        return False

    # Passo 2: Selecionar extrato
    print("📄 Agora, selecione seu extrato do Sicoob...")
    caminho_extrato = filedialog.askopenfilename(
        title="Selecione o arquivo de extrato do Sicoob",
        filetypes=[("Arquivos Excel/CSV", "*.xlsx;*.xls;*.csv")]
    )
    
    if not caminho_extrato:
        messagebox.showinfo("Cancelado", "Operação cancelada. Nenhum arquivo de extrato foi selecionado.")
        return False

    print(f"✅ Planilha criada: {os.path.basename(caminho_planilha_usuario)}")
    print(f"✅ Extrato selecionado: {os.path.basename(caminho_extrato)}")
    print()
    print("🔄 Iniciando processamento...")

    # Processa o extrato
    resultado = processar_extrato_individual(caminho_extrato, caminho_planilha_usuario, mostrar_detalhes=True)
    
    if resultado['sucesso']:
        # Mostrar exemplos dos lançamentos adicionados se houver
        if resultado['novos_lancamentos'] > 0:
            print(f"\n📋 Processamento concluído com sucesso!")
        
        # Pergunta se quer abrir a planilha
        abrir_planilha = messagebox.askyesno(
            "Sucesso!", 
            f"✅ Processamento concluído com sucesso!\n\n"
            f"📊 Transações processadas: {resultado['transacoes_processadas']}\n"
            f"💰 Débitos encontrados: {resultado['debitos_encontrados']}\n"
            f"➕ Novos lançamentos adicionados: {resultado['novos_lancamentos']}\n"
            f"🔍 Duplicatas ignoradas: {resultado['duplicatas_ignoradas']}\n"
            f"🎨 Formatação preservada!\n\n"
            f"📁 Planilha salva em:\n{os.path.basename(caminho_planilha_usuario)}\n\n"
            f"Deseja abrir a planilha agora?"
        )
        
        if abrir_planilha:
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(caminho_planilha_usuario)
                else:  # Linux/Mac
                    import subprocess
                    subprocess.call(['xdg-open', caminho_planilha_usuario])
            except Exception as e:
                print(f"Erro ao abrir planilha: {e}")
        
        return True
    else:
        messagebox.showerror("Erro", f"❌ Erro no processamento:\n\n{resultado.get('erro', 'Erro desconhecido')}")
        return False

def criar_menu_principal():
    """
    Cria menu principal com opções de processamento
    """
    # Janela principal
    root = tk.Tk()
    root.title("🏠 Automação Sicoob v2.1")
    root.geometry("600x450")
    root.resizable(False, False)
    
    # Centralizar janela
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    # Estilo
    root.configure(bg='#f5f5f5')
    
    # Frame principal
    frame_principal = tk.Frame(root, bg='#f5f5f5', padx=40, pady=30)
    frame_principal.pack(fill=tk.BOTH, expand=True)
    
    # Título
    label_titulo = tk.Label(frame_principal, text="🚀 AUTOMAÇÃO SICOOB", 
                           font=('Arial', 22, 'bold'), fg='#2E7D32', bg='#f5f5f5')
    label_titulo.pack(pady=(0, 10))
    
    label_subtitulo = tk.Label(frame_principal, text="Processamento Avançado de Extratos Bancários", 
                              font=('Arial', 12), fg='#666', bg='#f5f5f5')
    label_subtitulo.pack(pady=(0, 30))
    
    # Frame dos botões
    frame_botoes = tk.Frame(frame_principal, bg='#f5f5f5')
    frame_botoes.pack(expand=True)
    
    # Estilo dos botões
    btn_style = {
        'font': ('Arial', 12, 'bold'),
        'width': 35,
        'height': 2,
        'relief': 'flat',
        'cursor': 'hand2'
    }
    
    # Botão processamento único
    btn_unico = tk.Button(frame_botoes, text="📄 Processar Extrato Único", 
                         command=lambda: [root.destroy(), processar_extrato_unico()],
                         bg='#4CAF50', fg='white', **btn_style)
    btn_unico.pack(pady=8)
    
    # Descrição do botão único
    desc_unico = tk.Label(frame_botoes, 
                         text="Processa um extrato e cria uma planilha personalizada", 
                         font=('Arial', 10), fg='#666', bg='#f5f5f5')
    desc_unico.pack(pady=(0, 15))
    
    # Botão processamento múltiplo
    btn_multiplo = tk.Button(frame_botoes, text="📁 Processar Múltiplos Extratos", 
                            command=lambda: [root.destroy(), processar_multiplos_extratos()],
                            bg='#2196F3', fg='white', **btn_style)
    btn_multiplo.pack(pady=8)
    
    # Descrição do botão múltiplo
    desc_multiplo = tk.Label(frame_botoes, 
                            text="Processa vários extratos, cada um em uma planilha separada", 
                            font=('Arial', 10), fg='#666', bg='#f5f5f5')
    desc_multiplo.pack(pady=(0, 20))
    
    # Separador
    separador = tk.Frame(frame_botoes, height=1, bg='#ddd')
    separador.pack(fill=tk.X, pady=15)
    
    # Informações da versão
    info_versao = tk.Label(frame_botoes, 
                          text="Versão 2.1 • Template Embutido • Processamento Robusto", 
                          font=('Arial', 9), fg='#999', bg='#f5f5f5')
    info_versao.pack(pady=5)
    
    # Rodapé
    rodape = tk.Label(frame_principal, 
                     text="© 2025 Automação Sicoob • Desenvolvido com ❤️ para facilitar sua vida", 
                     font=('Arial', 9), fg='#999', bg='#f5f5f5')
    rodape.pack(side=tk.BOTTOM, pady=10)
    
    # Efeitos de hover nos botões
    def on_enter(e, cor_hover):
        e.widget.config(bg=cor_hover)
    
    def on_leave(e, cor_original):
        e.widget.config(bg=cor_original)
    
    btn_unico.bind("<Enter>", lambda e: on_enter(e, '#45a049'))
    btn_unico.bind("<Leave>", lambda e: on_leave(e, '#4CAF50'))
    
    btn_multiplo.bind("<Enter>", lambda e: on_enter(e, '#1976D2'))
    btn_multiplo.bind("<Leave>", lambda e: on_leave(e, '#2196F3'))
    
    return root

def main():
    """Função principal do programa"""
    print("🚀 Iniciando Automação Sicoob v2.1...")
    print("=" * 60)
    
    # Verificar dependências
    if not verificar_dependencias():
        print("❌ Dependências necessárias não estão instaladas!")
        input("Pressione ENTER para sair...")
        return
    
    try:
        # Criar e executar menu principal
        app = criar_menu_principal()
        app.mainloop()
        
    except Exception as e:
        print(f"❌ Erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        
    finally:
        print("\n" + "=" * 60)
        print("👋 Obrigado por usar a Automação Sicoob!")
        input("\nPressione ENTER para sair...")

if __name__ == "__main__":
    main()